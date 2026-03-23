import os
import re
import requests
from datetime import datetime

from openpyxl import load_workbook
from pdf2image import convert_from_path
import pytesseract
from pypdf import PdfReader


# =============================
# CONFIG
# =============================

EXCEL_FILE = "Cumpleaños.xlsx"

POPPLER_PATH = r"C:\Users\svara\Desktop\Release-24.08.0-0\poppler-24.08.0\Library\bin"
TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

PDF_FOLDER = "pdfs"

pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH

os.makedirs(PDF_FOLDER, exist_ok=True)


# =============================
# DRIVE
# =============================

def convertir_drive(url):
    match = re.search(r'/d/(.*?)/', url)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=download&id={file_id}"
    return url


# =============================
# DESCARGAR PDF
# =============================

def descargar_pdf(url, path):
    try:
        r = requests.get(url)
        if r.status_code == 200:
            with open(path, "wb") as f:
                f.write(r.content)
            return True
    except:
        pass
    return False


# =============================
# LEER PDF TEXTO
# =============================

def leer_pdf(path):
    texto = ""
    try:
        reader = PdfReader(path)
        for page in reader.pages:
            t = page.extract_text()
            if t:
                texto += t
    except:
        pass
    return texto


# =============================
# OCR
# =============================

def leer_ocr(path):
    texto = ""
    images = convert_from_path(path, poppler_path=POPPLER_PATH)

    for img in images:
        texto += pytesseract.image_to_string(img, lang="spa")

    return texto


# =============================
# LIMPIAR TEXTO
# =============================

def limpiar_texto(texto):

    texto = texto.lower()

    reemplazos = {
        "mar20": "marzo",
        "marz0": "marzo",
        "setiembre": "septiembre",
        "novienbre": "noviembre"
    }

    for k, v in reemplazos.items():
        texto = texto.replace(k, v)

    return texto


# =============================
# EXTRAER FECHAS
# =============================

def extraer_fechas(texto):

    texto = limpiar_texto(texto)

    meses = {
        "enero":1, "febrero":2, "marzo":3, "abril":4,
        "mayo":5, "junio":6, "julio":7, "agosto":8,
        "septiembre":9, "octubre":10, "noviembre":11, "diciembre":12
    }

    fechas = []

    matches = re.findall(r"(\d{1,2}) de ([a-z]+) de (\d{4})", texto)

    for d, m, y in matches:
        if m in meses:
            try:
                fechas.append(datetime(int(y), meses[m], int(d)))
            except:
                pass

    return fechas


# =============================
# DETECTAR FECHA FINAL
# =============================

def detectar_fecha(path):

    texto = leer_pdf(path)
    texto += leer_ocr(path)

    fechas = extraer_fechas(texto)

    if fechas:
        return max(fechas)  # 🔥 la última fecha = fin contrato

    return None


# =============================
# PROCESO PRINCIPAL
# =============================

def procesar():

    print("📄 LEYENDO CONTRATOS...\n")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in range(2, ws.max_row + 1):

        tipo = ws[f"A{row}"].value
        nombre = ws[f"B{row}"].value

        pdf_cell = ws[f"H{row}"]
        fecha_cell = ws[f"N{row}"]  # 🔥 COLUMNA N

        print(f"👤 {nombre}")

        # INDEFINIDO
        if tipo == "INDEFINIDO":
            fecha_cell.value = "∞"
            print("♾️ Indefinido\n")
            continue

        # SI YA TIENE FECHA → NO PROCESAR
        if fecha_cell.value not in (None, "", " "):
            print("✅ Ya tiene fecha\n")
            continue

        # SI NO TIENE PDF
        if not pdf_cell.hyperlink:
            print("❌ Sin PDF\n")
            continue

        link = pdf_cell.hyperlink.target
        url = convertir_drive(link)

        pdf_path = os.path.join(PDF_FOLDER, f"{row}.pdf")

        if not descargar_pdf(url, pdf_path):
            print("❌ Error descarga\n")
            continue

        fecha = detectar_fecha(pdf_path)

        if fecha:
            fecha_cell.value = fecha.strftime("%d-%m-%Y")  # 🔥 FORMATO QUE QUIERES
            print("📅 Fecha detectada:", fecha.strftime("%d-%m-%Y"), "\n")
        else:
            print("❌ No se detectó fecha\n")

    wb.save(EXCEL_FILE)

    print("✅ PROCESO TERMINADO")


procesar()