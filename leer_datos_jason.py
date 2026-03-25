# leer_datos_jason.py

import json
import os
from openpyxl import load_workbook
from datetime import datetime, date

EXCEL_FILE = "Cumpleaños.xlsx"
ESTADOS_FILE = "estados.json"


# =============================
# CARGAR ESTADOS JSON 🔥
# =============================
def cargar_estados():

    if not os.path.exists(ESTADOS_FILE):
        return {}

    try:
        with open(ESTADOS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}


def get_link(cell):

    if cell.hyperlink:
        return cell.hyperlink.target

    if isinstance(cell.value, str) and cell.value.startswith("http"):
        return cell.value

    if isinstance(cell.value, str) and "HYPERLINK" in cell.value.upper():
        try:
            link = cell.value.split('"')[1]
            return link
        except:
            return ""

    return ""


def calcular_dias_contrato(fecha):
    hoy = datetime.today()
    return (fecha - hoy).days + 1


def calcular_dias_cumple(cumple):

    if not cumple:
        return ""

    try:
        hoy = date.today()

        cumple_este_año = date(hoy.year, cumple.month, cumple.day)

        if cumple_este_año < hoy:
            cumple_este_año = date(hoy.year + 1, cumple.month, cumple.day)

        return (cumple_este_año - hoy).days

    except:
        return ""


def generar_json():

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    estados_guardados = cargar_estados()  # 🔥 NUEVO

    datos = []

    for row in range(2, ws.max_row + 1):

        nombre = ws[f"B{row}"].value
        tipo = ws[f"A{row}"].value
        rut = ws[f"E{row}"].value
        cumple = ws[f"C{row}"].value
        fecha_raw = ws[f"N{row}"].value

        rut_str = str(rut).strip() if rut else ""

        # =============================
        # 🔥 ESTADO INTELIGENTE
        # =============================

        # 1. Buscar en estados.json
        estado = estados_guardados.get(rut_str)

        # 2. Si no existe → usar Excel (columna O)
        if not estado:
            estado_excel = ws[f"O{row}"].value

            if estado_excel:
                estado = str(estado_excel).strip().upper()
            else:
                estado = "ACTIVO"

        # 3. Seguridad
        if estado not in ["ACTIVO", "INACTIVO"]:
            estado = "ACTIVO"

        # =============================

        if fecha_raw == "∞" or tipo == "INDEFINIDO":

            dias = "INDEFINIDO"
            fecha_termino = "∞"

        else:

            try:

                if isinstance(fecha_raw, datetime):
                    fecha = fecha_raw
                else:
                    fecha = datetime.strptime(str(fecha_raw), "%d-%m-%Y")

                dias = calcular_dias_contrato(fecha)
                fecha_termino = fecha.strftime("%d-%m-%Y")

            except:

                dias = ""
                fecha_termino = None

        persona = {

            "id": row,
            "nombre": nombre,
            "tipo": tipo,
            "dias": dias,
            "fecha_termino": fecha_termino,
            "rut": rut_str,
            "estado": estado,

            "cumple": cumple.strftime("%d/%m/%Y") if isinstance(cumple, datetime) else "",
            "dias_cumple": calcular_dias_cumple(cumple),

            "pdfs": {

                "ci": get_link(ws[f"F{row}"]),
                "contrato": get_link(ws[f"H{row}"]),
                "psi": get_link(ws[f"J{row}"]),
                "lc": get_link(ws[f"L{row}"]),
                "informe": get_link(ws[f"M{row}"])

            }

        }

        datos.append(persona)

    with open("datos.json", "w", encoding="utf-8") as f:
        json.dump(datos, f, indent=4, ensure_ascii=False)

    print("✅ datos.json generado correctamente")


if __name__ == "__main__":
    generar_json()
