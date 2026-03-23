import os
import pandas as pd
from openpyxl import load_workbook

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# =========================
# CONFIG
# =========================

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']

excel_path = "Cumpleaños.xlsx"

FOLDER_ID = "1MAleZ8QRbl7ldXYlORErhGQP_ptnYIxq"

# =========================
# AUTENTICACION GOOGLE
# =========================

creds = None

if os.path.exists("token.json"):
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)

if not creds:

    flow = InstalledAppFlow.from_client_secrets_file(
        "credenciales.json", SCOPES)

    creds = flow.run_local_server(port=0)

    with open("token.json", "w") as token:
        token.write(creds.to_json())

service = build("drive", "v3", credentials=creds)

print("Conectado a Google Drive")

# =========================
# CARGAR EXCEL
# =========================

df = pd.read_excel(excel_path)

wb = load_workbook(excel_path)
ws = wb.active

col_pdf = {
    "PDF PSI": 10,
    "PDF LC": 12,
    "PDF INFORME": 13,
    "PDF CT": 8,
    "PDF CI": 6
}

# =========================
# INVERTIR NOMBRE
# =========================

def invertir_nombre(nombre):

    partes = nombre.strip().split()

    if len(partes) >= 2:
        nombre = partes[0]
        apellido = partes[-1]
        return f"{apellido} {nombre}".upper()

    return nombre.upper()

# =========================
# BUSCAR CARPETA
# =========================

def buscar_carpeta(nombre):

    query = f"name contains '{nombre}' and '{FOLDER_ID}' in parents and mimeType='application/vnd.google-apps.folder'"

    results = service.files().list(
        q=query,
        fields="files(id,name)"
    ).execute()

    files = results.get("files", [])

    if files:
        return files[0]["id"]

    return None

# =========================
# OBTENER PDFS
# =========================

def obtener_pdfs(folder_id):

    query = f"'{folder_id}' in parents and mimeType='application/pdf' and trashed=false"

    results = service.files().list(
        q=query,
        fields="files(id,name)"
    ).execute()

    return results.get("files", [])

# =========================
# RECORRER EXCEL
# =========================

for i, row in df.iterrows():

    nombre_excel = str(row["NOMBRE"]).upper()

    nombre_busqueda = invertir_nombre(nombre_excel)

    print("Buscando carpeta:", nombre_busqueda)

    folder_id = buscar_carpeta(nombre_busqueda)

    if not folder_id:
        print("Carpeta no encontrada")
        continue

    archivos = obtener_pdfs(folder_id)

    fila_excel = i + 2

    for archivo in archivos:

        nombre_archivo = archivo["name"].upper()
        file_id = archivo["id"]

        link = f"https://drive.google.com/file/d/{file_id}/view"

        # =========================
        # PSI
        # =========================

        if "PSICOSENSOTECNICO" in nombre_archivo:

            celda = ws.cell(row=fila_excel, column=col_pdf["PDF PSI"])

            if celda.value is None:
                celda.value = "Abrir PDF"
                celda.hyperlink = link
                celda.style = "Hyperlink"

        # =========================
        # LICENCIA
        # =========================

        elif nombre_archivo.startswith("LC"):

            celda = ws.cell(row=fila_excel, column=col_pdf["PDF LC"])

            if celda.value is None:
                celda.value = "Abrir PDF"
                celda.hyperlink = link
                celda.style = "Hyperlink"

        # =========================
        # INFORME
        # =========================

        elif "INFORME" in nombre_archivo:

            celda = ws.cell(row=fila_excel, column=col_pdf["PDF INFORME"])

            if celda.value is None:
                celda.value = "Abrir PDF"
                celda.hyperlink = link
                celda.style = "Hyperlink"

        # =========================
        # CONTRATO
        # =========================

        elif nombre_archivo.startswith("CT"):

            celda = ws.cell(row=fila_excel, column=col_pdf["PDF CT"])

            if celda.value is None:
                celda.value = "Abrir PDF"
                celda.hyperlink = link
                celda.style = "Hyperlink"

        # =========================
        # CEDULA IDENTIDAD
        # =========================

        elif "CI" in nombre_archivo:

            celda = ws.cell(row=fila_excel, column=col_pdf["PDF CI"])

            if celda.value is None:
                celda.value = "Abrir PDF"
                celda.hyperlink = link
                celda.style = "Hyperlink"

# =========================
# GUARDAR EXCEL
# =========================

import sys

wb.save(excel_path)

print("Excel actualizado correctamente")

sys.exit(0)