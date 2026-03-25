from flask import Flask, request, redirect, session, send_from_directory, send_file
import json
import os
from datetime import timedelta
from leer_datos_jason import generar_json
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "Empresacoldcontrolcontactocoldcontrol"

app.permanent_session_lifetime = timedelta(days=30)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ESTADOS_FILE = os.path.join(BASE_DIR, "estados.json")


# =============================
# CARGAR USUARIOS
# =============================
def cargar_usuarios():
    with open(os.path.join(BASE_DIR, "usuarios.json"), "r", encoding="utf-8") as f:
        return json.load(f)["usuarios"]


# =============================
# GUARDAR ESTADOS 🔥 NUEVO
# =============================
def guardar_estado_json(rut, estado):

    try:

        # Crear archivo si no existe
        if not os.path.exists(ESTADOS_FILE):
            with open(ESTADOS_FILE, "w", encoding="utf-8") as f:
                json.dump({}, f)

        # Leer estados actuales
        with open(ESTADOS_FILE, "r", encoding="utf-8") as f:
            estados = json.load(f)

        # Actualizar estado
        estados[rut] = estado

        # Guardar
        with open(ESTADOS_FILE, "w", encoding="utf-8") as f:
            json.dump(estados, f, indent=4, ensure_ascii=False)

    except Exception as e:
        print("Error guardando estados.json:", e)


# =============================
# LOGIN
# =============================
@app.route("/", methods=["GET", "POST"])
def login():

    if "user" in session:
        return redirect("/index.html")

    if request.method == "POST":

        user = request.form["usuario"].lower()
        password = request.form["password"]
        recordar = request.form.get("recordar")

        usuarios = cargar_usuarios()

        for u in usuarios:
            if u["usuario"] == user and u["password"] == password:
                session["user"] = user
                session.permanent = bool(recordar)
                return redirect("/index.html")

        return redirect("/?error=1")

    return send_from_directory(BASE_DIR, "login.html")


# =============================
# INDEX
# =============================
@app.route("/index.html")
def index():
    if "user" not in session:
        return redirect("/")
    return send_from_directory(BASE_DIR, "index.html")


# =============================
# DETALLE
# =============================
@app.route("/detalle.html")
def detalle():
    if "user" not in session:
        return redirect("/")
    return send_from_directory(BASE_DIR, "detalle.html")


# =============================
# DATOS JSON
# =============================
@app.route("/datos.json")
def datos():

    if "user" not in session:
        return {"error": "no autorizado"}, 403

    generar_json()

    return send_file(os.path.join(BASE_DIR, "datos.json"))


# =============================
# CAMBIAR ESTADO 🔥 MEJORADO
# =============================
@app.route("/desactivar/<rut>", methods=["POST"])
def desactivar(rut):

    if "user" not in session:
        return {"error": "no autorizado"}, 403

    try:

        data = request.get_json()
        nuevo_estado = data.get("estado", "INACTIVO")

        archivo_excel = os.path.join(BASE_DIR, "Cumpleaños.xlsx")

        wb = load_workbook(archivo_excel)
        ws = wb.active

        for row in range(2, ws.max_row + 1):

            rut_excel = str(ws[f"E{row}"].value)

            if rut_excel == rut:

                # 🔹 Guardar en Excel
                ws[f"O{row}"] = nuevo_estado

                break

        wb.save(archivo_excel)

        # 🔥 GUARDAR EN JSON (PERSISTENTE)
        guardar_estado_json(rut, nuevo_estado)

        # 🔄 Regenerar datos
        generar_json()

        return {"ok": True}

    except Exception as e:

        print("Error:", e)

        return {"ok": False}


# =============================
# LOGOUT
# =============================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# =============================
# RUN SERVER
# =============================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
